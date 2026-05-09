import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

HEADER_FONT = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1A3A6A")
CAT_FILL = {
    "大変更": PatternFill("solid", fgColor="FFE5E5"),
    "業務ロジック": PatternFill("solid", fgColor="FFF2CC"),
    "マスタ": PatternFill("solid", fgColor="E2F0D9"),
    "REPORT3 / 領収書": PatternFill("solid", fgColor="DEEBF7"),
    "その他": PatternFill("solid", fgColor="EDEDED"),
}
ANSWER_FILL = PatternFill("solid", fgColor="FFFBEA")
THIN = Side(border_style="thin", color="888888")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
JP_FONT = Font(name="Yu Gothic", size=10)
JP_FONT_BOLD = Font(name="Yu Gothic", size=10, bold=True)

# シート1: 質問一覧
ws = wb.active
ws.title = "質問一覧"
for i, w in enumerate([6, 18, 6, 50, 30, 40], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.merge_cells("A1:F1")
ws["A1"] = "さくら株式会社 業務管理システム 仕様確認"
ws["A1"].font = Font(name="Yu Gothic", size=16, bold=True, color="1A3A6A")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:F2")
ws["A2"] = "下記の質問の「回答」列(F列・クリーム色)にご記入ください。Q8(REPORT3 工種分類)は別シートに記入欄をご用意しております。"
ws["A2"].font = Font(name="Yu Gothic", size=10, color="666666")
ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

for i, h in enumerate(["Q番号", "カテゴリ", "枝番", "質問内容", "選択肢 / 補足", "回答(ここにご記入ください)"], 1):
    c = ws.cell(row=4, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = WRAP_C
    c.border = BORDER
ws.row_dimensions[4].height = 26

questions = [
    ("Q1", "大変更", "1-1", "受け付けるファイル形式は Excel + PDF + Word + 画像 すべてで問題ございませんでしょうか。1ファイルあたりの容量上限はいかがいたしますでしょうか。", "形式: 全て可 / 一部のみ\n容量上限: ___ MB"),
    ("Q1", "大変更", "1-2", "元請さまごとにフォーマットが異なるとのことでしたが、システム側でテンプレートを管理する仕組みが必要でしょうか。それともアップロード機能のみで十分でしょうか。", "A. テンプレート管理あり\nB. アップロードのみ"),
    ("Q1", "大変更", "1-3", "入構人員情報(資格/住所/連絡先)は USER マスタから自動引用する想定でよろしいでしょうか。", "Yes / No"),
    ("Q1", "大変更", "1-4", "上記のため、住所と連絡先の項目を USER マスタに新規追加してよろしいでしょうか。", "Yes / No"),
    ("Q1", "大変更", "1-5", "元請さまと協力会社さまの両方それぞれに別個に提出するのでしょうか。それとも1セットで両方を兼ねるのでしょうか。", "A. 別個 / B. 1セット"),
    ("Q2", "業務ロジック", "2-1", "「受注済み」に該当するステータスは現状の見積書の「accepted(受注)」で合致するという理解でよろしいでしょうか。", "Yes / No"),
    ("Q2", "業務ロジック", "2-2", "複数の見積が1案件に紐づく場合、どの見積を概況表に反映するのが適切でしょうか。", "A. 最新の受注見積のみ\nB. 全受注見積の合計\nC. 1案件=1受注見積ルール"),
    ("Q3", "業務ロジック", "3-1", "工事経歴書の出力形式は PDF / Excel / 両方 のうち、どれが望ましいでしょうか。", "PDF / Excel / 両方"),
    ("Q3", "業務ロジック", "3-2", "既存のテンプレート(建業許可で使用される様式)がございましたら、ご共有いただけますでしょうか。", "あり(添付) / なし"),
    ("Q3", "業務ロジック", "3-3", "1案件で複数回請求書を発行された場合、1案件あたり1通の工事経歴書を作成するのか、請求書ごとに作成するのか、いずれがよろしいでしょうか。", "A. 1案件1通\nB. 請求書ごと"),
    ("Q3", "業務ロジック", "3-4", "和暦の表記は「令和7年5月7日」形式でよろしいでしょうか。", "Yes / No(別表記指定)"),
    ("Q4", "業務ロジック", "4-1a", "経費管理表 - 見積金額 の入力者は?", "自動 / 事務手入力 / 案件責任者"),
    ("Q4", "業務ロジック", "4-1b", "経費管理表 - 仕入(材料)金額 の入力者は?", "自動 / 事務手入力 / 案件責任者"),
    ("Q4", "業務ロジック", "4-1c", "経費管理表 - リース金額 の入力者は?", "自動 / 事務手入力 / 案件責任者"),
    ("Q4", "業務ロジック", "4-1d", "経費管理表 - 協力会社(人工/残業/金額) の入力者は?", "自動 / 事務手入力 / 案件責任者"),
    ("Q4", "業務ロジック", "4-1e", "経費管理表 - 諸経費 の入力者は?", "自動 / 事務手入力 / 案件責任者"),
    ("Q4", "業務ロジック", "4-1f", "経費管理表 - 車両交通費 の入力者は?", "自動(領収書) / 手入力"),
    ("Q4", "業務ロジック", "4-1g", "経費管理表 - 人件費(人工/残業/出張費) の入力者は?", "自動(REPORT3) / 手入力"),
    ("Q4", "業務ロジック", "4-1h", "経費管理表 - 宿泊費(レオパレス/ホテル) の入力者は?", "自動(領収書) / 手入力"),
    ("Q4", "業務ロジック", "4-2", "見積金額は見積書から自動取得する形でよろしいでしょうか。", "Yes / No"),
    ("Q4", "業務ロジック", "4-3", "協力会社の金額入力方法 — 別途請求書を入力する仕組みが必要でしょうか。", "A. 必要(機能追加)\nB. 不要(手入力で対応)"),
    ("Q5", "マスタ", "5-1", "「職人スキル」の評価軸はどのようなものを想定されていますでしょうか。", "数値スコア / 自己申告 / 上司評価 / 複合 / その他"),
    ("Q5", "マスタ", "5-2", "「人事評価データ」の閲覧権限はどなたまでとされますでしょうか。", "本人のみ / 直属の上司まで / 経営層のみ"),
    ("Q5", "マスタ", "5-3", "「経験現場履歴」は REPORT3 から自動構築する形でよろしいでしょうか。", "Yes / No"),
    ("Q6", "マスタ", "6-1", "「松竹梅」の判定基準はどのように設定されますでしょうか。", "案件作成時に手動 / 案件規模で自動 / その他"),
    ("Q6", "マスタ", "6-2", "役職別単価は組織図の役職テーブルから決定する形でよろしいでしょうか(例:職長=標準の120%)。", "Yes / No(別ロジック指定)"),
    ("Q6", "マスタ", "6-3", "単価決定のタイミングは、見積入力時に自動計算まで実施 か、推奨単価を表示してご担当者が調整 か、どちらが望ましいでしょうか。", "A. 自動計算\nB. 推奨表示+手動調整"),
    ("Q7", "マスタ", "7-1", "Excel のテンプレート様式は社長様にてご用意いただけますでしょうか。それとも当方で雛形をご提案いたしましょうか。", "A. 社長提供\nB. 当方で提案"),
    ("Q7", "マスタ", "7-2", "「公開/非公開」のフラグは組織図全体に対するものか、個別の更新ごとに切り替えるものか、いずれをご想定でしょうか。", "A. 全体 / B. 個別"),
    ("Q7", "マスタ", "7-3", "「役職履歴」の粒度は日単位 / 月単位 / 任意の日付指定 のどれが業務上扱いやすいでしょうか。", "日単位 / 月単位 / 任意日付"),
    ("Q8", "REPORT3 / 領収書", "8-*", "REPORT3 の工種分類 4 階層化について。詳細は別シート「Q8-中分類」「Q8-小分類」「Q8-組合せ」「Q8-反映方針」をご参照ください。", "別シートに記入"),
    ("Q9", "REPORT3 / 領収書", "9-1", "標準勤務時間の前提を 8:00〜17:00 / 休憩1時間(計8時間)としてよろしいでしょうか。", "Yes / No(別時間指定)"),
    ("Q9", "REPORT3 / 領収書", "9-2", "夜勤の判定はどの時間帯で行うのが適切でしょうか(例:勤務開始時刻が22:00以降を夜勤)。", "境界時刻: ___ 時 以降"),
    ("Q9", "REPORT3 / 領収書", "9-3", "夜勤分は除外とするか、別カウントとするか、いずれがよろしいでしょうか。", "A. 除外\nB. 別カウント(夜勤手当対象)"),
    ("Q9", "REPORT3 / 領収書", "9-4", "スケジュール機能(配車表・人員配置)の先行実装が必要との理解でよろしいでしょうか。", "Yes(先行実装) / No(別途簡易実装)"),
    ("Q10", "REPORT3 / 領収書", "10-1", "「精算の有無」とは、会社負担で立替済→精算が必要 か 個人負担(精算不要)の2択でよろしいでしょうか。", "Yes / No(別の意味)"),
    ("Q10", "REPORT3 / 領収書", "10-2", "カテゴリ選択肢(消耗品/食事代/交通費/宿泊費/その他)で過不足ございませんでしょうか。追加項目があればご記入ください。", "追加項目: ___________"),
    ("Q10", "REPORT3 / 領収書", "10-3", "1件あたりの金額上限(例:5万円超は別途承認フロー)を設定する必要はございますでしょうか。", "A. 不要 / B. 必要(上限: ___ 円)"),
    ("Q10", "REPORT3 / 領収書", "10-4", "通知先となる「総務課」は既存「事務」ロール内の特定スタッフ(フラグ)としてよろしいでしょうか。それとも新規ロールに分離しますか。", "A. 既存ロール+フラグ\nB. 新規ロール分離"),
    ("Q11", "その他", "11-1", "Eight 連携の方向性。Eight→顧客マスタへの取込のみで十分でしょうか。書き出しも必要でしょうか。", "A. 取込のみ\nB. 双方向"),
    ("Q11", "その他", "11-2", "Eight の API キーは社長様側でご用意いただける想定でよろしいでしょうか。", "Yes / No(検討中)"),
    ("Q12", "その他", "12-1", "事務ロールを「総務課」「施工管理課」の2つに分けるか、ロールは「事務」のままで部署マスタで識別するか、いずれがよろしいでしょうか。", "A. ロール分割\nB. 部署マスタで識別"),
]

for i, (q, cat, sub, qtext, options) in enumerate(questions, 5):
    cells = [
        ws.cell(row=i, column=1, value=q),
        ws.cell(row=i, column=2, value=cat),
        ws.cell(row=i, column=3, value=sub),
        ws.cell(row=i, column=4, value=qtext),
        ws.cell(row=i, column=5, value=options),
        ws.cell(row=i, column=6, value=""),
    ]
    for c in cells[:5]:
        c.font = JP_FONT
        c.alignment = WRAP
        c.border = BORDER
    cells[0].font = JP_FONT_BOLD
    cells[0].alignment = WRAP_C
    cells[2].alignment = WRAP_C
    if cat in CAT_FILL:
        cells[1].fill = CAT_FILL[cat]
    cells[5].fill = ANSWER_FILL
    cells[5].border = BORDER
    cells[5].alignment = WRAP
    cells[5].font = JP_FONT
    qlen = len(qtext) + len(options or "")
    ws.row_dimensions[i].height = max(40, min(110, qlen * 0.7))


def add_q8_chubun():
    ws2 = wb.create_sheet("Q8-中分類")
    for i, w in enumerate([16, 8, 35], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Q8 中分類リスト(大分類ごとに記入)"
    ws2["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24
    ws2.merge_cells("A2:C2")
    ws2["A2"] = "「大分類」を選択した時に表示される「中分類」のリストをご記入ください。「配管」は記載済(参考)。1セルに1項目をご記入ください。"
    ws2["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws2["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws2.row_dimensions[2].height = 30
    for i, h in enumerate(["大分類", "No.", "中分類(ご記入欄)"], 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_C
        c.border = BORDER
    ws2.row_dimensions[4].height = 22
    chubun = {
        "配管": ["冷媒", "冷温水", "冷却水", "排水", "食品プロセス", "ケミカルプロセス", "エネルギープロセス", "蒸気", "エアー"],
        "鳶": [""] * 10,
        "保温": [""] * 10,
        "溶接": [""] * 10,
        "雑工事": [""] * 10,
    }
    row = 5
    for daibun, lst in chubun.items():
        start = row
        for j, item in enumerate(lst, 1):
            c1 = ws2.cell(row=row, column=1, value=daibun if j == 1 else "")
            c2 = ws2.cell(row=row, column=2, value=j)
            c3 = ws2.cell(row=row, column=3, value=item)
            c1.font = JP_FONT_BOLD
            c1.alignment = Alignment(vertical="center", horizontal="center")
            c2.font = JP_FONT
            c2.alignment = WRAP_C
            c3.font = JP_FONT
            c3.alignment = WRAP
            for c in [c1, c2, c3]:
                c.border = BORDER
            if daibun == "配管":
                c3.fill = PatternFill("solid", fgColor="EEEEEE")
            else:
                c3.fill = ANSWER_FILL
            ws2.row_dimensions[row].height = 22
            row += 1
        if len(lst) > 1:
            ws2.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        row += 1


def add_q8_shoubun():
    ws3 = wb.create_sheet("Q8-小分類")
    for i, w in enumerate([18, 8, 35], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Q8 小分類リスト(中分類ごとに記入)"
    ws3["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24
    ws3.merge_cells("A2:C2")
    ws3["A2"] = "「中分類」を選択した時に表示される「小分類」のリストをご記入ください。「排水」は記載済(参考)。1セルに1項目をご記入ください。"
    ws3["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws3["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws3.row_dimensions[2].height = 30
    for i, h in enumerate(["中分類", "No.", "小分類(ご記入欄)"], 1):
        c = ws3.cell(row=4, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_C
        c.border = BORDER
    ws3.row_dimensions[4].height = 22
    shoubun = {
        "冷媒": [""] * 10,
        "冷温水": [""] * 10,
        "冷却水": [""] * 10,
        "排水": ["塩ビ/のり付け", "塩ビ/溶接", "塩ビ/加工管・フランジ止め", "SGPW/溶接", "SGPW/ネジ", "SGPW/加工管・フランジ閉め", "SUS", "SGP_黒"],
        "食品プロセス": [""] * 10,
        "ケミカルプロセス": [""] * 10,
        "エネルギープロセス": [""] * 10,
        "蒸気": [""] * 10,
        "エアー": [""] * 10,
    }
    row = 5
    for chubun_name, lst in shoubun.items():
        start = row
        for j, item in enumerate(lst, 1):
            c1 = ws3.cell(row=row, column=1, value=chubun_name if j == 1 else "")
            c2 = ws3.cell(row=row, column=2, value=j)
            c3 = ws3.cell(row=row, column=3, value=item)
            c1.font = JP_FONT_BOLD
            c1.alignment = Alignment(vertical="center", horizontal="center")
            c2.font = JP_FONT
            c2.alignment = WRAP_C
            c3.font = JP_FONT
            c3.alignment = WRAP
            for c in [c1, c2, c3]:
                c.border = BORDER
            if chubun_name == "排水":
                c3.fill = PatternFill("solid", fgColor="EEEEEE")
            else:
                c3.fill = ANSWER_FILL
            ws3.row_dimensions[row].height = 22
            row += 1
        if len(lst) > 1:
            ws3.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        row += 1


def add_q8_combo():
    ws4 = wb.create_sheet("Q8-組合せ")
    ws4.column_dimensions["A"].width = 22
    for i in range(2, 7):
        ws4.column_dimensions[get_column_letter(i)].width = 14
    ws4.merge_cells("A1:F1")
    ws4["A1"] = "Q8 特大分類 × 大分類 の組み合わせ制限"
    ws4["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24
    ws4.merge_cells("A2:F2")
    ws4["A2"] = "業務上発生しうる組合せに「○」、発生しない/制限する組合せに「×」をご記入ください。空白は判断保留。"
    ws4["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws4["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws4.row_dimensions[2].height = 30
    ws4.cell(row=4, column=1, value="特大分類 \\ 大分類")
    for i, d in enumerate(["配管", "鳶", "保温", "溶接", "雑工事"], 2):
        ws4.cell(row=4, column=i, value=d)
    for i in range(1, 7):
        c = ws4.cell(row=4, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_C
        c.border = BORDER
    ws4.row_dimensions[4].height = 24
    for r, t in enumerate(["冷凍冷蔵", "ケミカルプラント", "大型半導体", "食品工場"], 5):
        c = ws4.cell(row=r, column=1, value=t)
        c.font = JP_FONT_BOLD
        c.alignment = Alignment(vertical="center", horizontal="center")
        c.fill = PatternFill("solid", fgColor="DDEEFF")
        c.border = BORDER
        for col in range(2, 7):
            cc = ws4.cell(row=r, column=col, value="")
            cc.fill = ANSWER_FILL
            cc.border = BORDER
            cc.alignment = WRAP_C
            cc.font = JP_FONT
        ws4.row_dimensions[r].height = 24


def add_q8_policy():
    ws5 = wb.create_sheet("Q8-反映方針")
    ws5.column_dimensions["A"].width = 50
    ws5.column_dimensions["B"].width = 30
    ws5.merge_cells("A1:B1")
    ws5["A1"] = "Q8 既存システムへの反映方針"
    ws5["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 24
    ws5.cell(row=3, column=1, value="質問")
    ws5.cell(row=3, column=2, value="ご回答")
    for i in range(1, 3):
        c = ws5.cell(row=3, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = WRAP_C
        c.border = BORDER
    ws5.row_dimensions[3].height = 22
    q4 = [
        "現状システムには3階層のシードデータ(33件)が登録されています。新4階層体系の導入方針はいかがいたしますでしょうか。\n\n選択肢: A. 既存3階層を完全削除→新4階層に入れ替え / B. 既存に追記する形で4階層を加える",
        "過去日報の移行対応は不要との認識でよろしいでしょうか。(現時点で本番データはない想定です)\n\n選択肢: Yes(不要) / No(必要)",
    ]
    for i, q in enumerate(q4, 4):
        ws5.cell(row=i, column=1, value=q)
        ws5.cell(row=i, column=2, value="")
        for j in range(1, 3):
            c = ws5.cell(row=i, column=j)
            c.font = JP_FONT
            c.alignment = WRAP
            c.border = BORDER
        ws5.cell(row=i, column=2).fill = ANSWER_FILL
        ws5.row_dimensions[i].height = 100


def add_legend():
    ws6 = wb.create_sheet("凡例", 0)
    ws6.column_dimensions["A"].width = 18
    ws6.column_dimensions["B"].width = 70
    ws6["A1"] = "本ファイルの構成"
    ws6["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws6.row_dimensions[1].height = 24
    rows = [
        ("シート", "内容"),
        ("質問一覧", "全12質問のサマリー。F列の回答欄(クリーム色)にご記入ください。"),
        ("Q8-中分類", "REPORT3 工種分類の中分類リストを大分類別に記入。"),
        ("Q8-小分類", "REPORT3 工種分類の小分類リストを中分類別に記入。"),
        ("Q8-組合せ", "特大分類×大分類のマトリックスに ○/× を記入。"),
        ("Q8-反映方針", "既存シードデータの取り扱いをご選択。"),
        ("", ""),
        ("色凡例", ""),
        ("クリーム色", "ご記入欄(自由にご記入ください)"),
        ("グレー", "記載済(参考表示、編集不要)"),
        ("色付ヘッダー", "カテゴリ別の色分け(赤=大変更、黄=業務ロジック、緑=マスタ、青=REPORT3、灰=その他)"),
    ]
    for i, (a, b) in enumerate(rows, 3):
        ca = ws6.cell(row=i, column=1, value=a)
        cb = ws6.cell(row=i, column=2, value=b)
        ca.font = JP_FONT_BOLD if i == 3 else JP_FONT
        cb.font = JP_FONT
        if i == 3:
            for col in [1, 2]:
                ws6.cell(row=i, column=col).fill = HEADER_FILL
                ws6.cell(row=i, column=col).font = HEADER_FONT
        cb.alignment = WRAP


add_q8_chubun()
add_q8_shoubun()
add_q8_combo()
add_q8_policy()
add_legend()

out_path = r"C:\Users\liim1\Desktop\SAKURAOS_仕様確認_質問票.xlsx"
wb.save(out_path)
print("Saved:", out_path)
