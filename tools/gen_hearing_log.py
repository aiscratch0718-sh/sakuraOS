"""
SAKURAOS ヒアリング項目 累積管理 Excel 生成スクリプト

このスクリプトは:
- これまでの全質問(Q1〜Q12 + 新規 Q13〜Q15)
- 回答(秋元様分のみ反映、Q8 / Q6-1 / Q6-3 は社長対応中)
- ステータス(未回答 / 社長確認待ち / 回答済 / 実装済)
- 起票日・回答日・担当
を 1 つの Excel にまとめる。今後の追加質問もこのファイルに追記していく運用。
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = openpyxl.Workbook()

HF = Font(name="Yu Gothic", size=11, bold=True, color="FFFFFF")
HFL = PatternFill("solid", fgColor="1A3A6A")
JF = Font(name="Yu Gothic", size=10)
JFB = Font(name="Yu Gothic", size=10, bold=True)
TH = Side(border_style="thin", color="888888")
BD = Border(left=TH, right=TH, top=TH, bottom=TH)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")

# ステータス別の塗り
STATUS_FILL = {
    "未回答": PatternFill("solid", fgColor="FFE5E5"),
    "社長確認待ち": PatternFill("solid", fgColor="FFF2CC"),
    "回答済": PatternFill("solid", fgColor="E2F0D9"),
    "実装済": PatternFill("solid", fgColor="DEEBF7"),
    "保留": PatternFill("solid", fgColor="EDEDED"),
}

CAT_FILL = {
    "大変更": PatternFill("solid", fgColor="FFE5E5"),
    "業務ロジック": PatternFill("solid", fgColor="FFF2CC"),
    "マスタ": PatternFill("solid", fgColor="E2F0D9"),
    "REPORT3 / 領収書": PatternFill("solid", fgColor="DEEBF7"),
    "その他": PatternFill("solid", fgColor="EDEDED"),
    "UI/UX": PatternFill("solid", fgColor="F0E5FF"),
}

ANSWER_FILL = PatternFill("solid", fgColor="FFFBEA")

# ========================================
# シート: 凡例
# ========================================
ws_idx = wb.active
ws_idx.title = "凡例"
ws_idx.column_dimensions["A"].width = 22
ws_idx.column_dimensions["B"].width = 80

ws_idx["A1"] = "SAKURAOS ヒアリング項目 累積管理"
ws_idx["A1"].font = Font(name="Yu Gothic", size=16, bold=True, color="1A3A6A")
ws_idx.row_dimensions[1].height = 28

ws_idx.merge_cells("A2:B2")
ws_idx["A2"] = (
    "クライアントへの確認事項を 1 ファイルで管理します。"
    "新たに質問が出たら『ヒアリング項目』シートの末尾に行を追加してください。"
)
ws_idx["A2"].font = Font(name="Yu Gothic", size=10, color="666666")
ws_idx["A2"].alignment = WRAP
ws_idx.row_dimensions[2].height = 36

intro = [
    ("シート構成", ""),
    ("ヒアリング項目", "全質問の一覧。ステータス(未回答/社長確認待ち/回答済/実装済)で管理。"),
    ("Q8-中分類", "REPORT3 工種分類: 大分類別の中分類リスト記入欄。社長対応中。"),
    ("Q8-小分類", "REPORT3 工種分類: 中分類別の小分類リスト記入欄。社長対応中。"),
    ("Q8-組合せ", "特大分類×大分類の組合せ制限マトリックス。社長対応中。"),
    ("Q8-反映方針", "既存シードデータの取扱い方針。社長対応中。"),
    ("", ""),
    ("ステータス凡例", ""),
    ("未回答(赤)", "クライアントへの未送付・回答待ち"),
    ("社長確認待ち(黄)", "秋元様回答済だが、社長確認が必要な項目"),
    ("回答済(緑)", "クライアント回答完了、実装着手待ち or 実装中"),
    ("実装済(青)", "回答内容に基づき実装が完了し、本番反映済み"),
    ("保留(灰)", "判断保留・再検討中・優先度低"),
    ("", ""),
    ("カテゴリ凡例", ""),
    ("大変更", "図面削除・安全書類など、設計の根幹に関わる項目"),
    ("業務ロジック", "見積→概況・経費管理表など、業務フローのロジック"),
    ("マスタ", "USER / QUAL / ORG / 部署など、マスタの構造・運用"),
    ("REPORT3 / 領収書", "日報・領収書・経費・夜勤判定など"),
    ("UI/UX", "メニュー表示形式・画面レイアウトなど"),
    ("その他", "外部連携・ロール・運用全般"),
    ("", ""),
    ("運用ルール", ""),
    ("追加方法", "『ヒアリング項目』末尾に行を追加。Q番号は連番(Q16, Q17...)"),
    ("送付タイミング", "未回答が一定数貯まったら一括送付、もしくは緊急度で随時"),
    ("更新履歴", "末尾の『更新履歴』シートに変更日と内容を追記"),
]
for i, (a, b) in enumerate(intro, 4):
    ca = ws_idx.cell(row=i, column=1, value=a)
    cb = ws_idx.cell(row=i, column=2, value=b)
    ca.font = JFB if (a in ("シート構成", "ステータス凡例", "カテゴリ凡例", "運用ルール")) else JF
    cb.font = JF
    cb.alignment = WRAP
    if a in ("シート構成", "ステータス凡例", "カテゴリ凡例", "運用ルール"):
        for col in [1, 2]:
            c = ws_idx.cell(row=i, column=col)
            c.fill = HFL
            c.font = HF
    elif a in STATUS_FILL:
        ca.fill = STATUS_FILL[a.split("(")[0]]
    elif a in CAT_FILL:
        ca.fill = CAT_FILL[a]

# ========================================
# シート: ヒアリング項目(メイン)
# ========================================
ws = wb.create_sheet("ヒアリング項目")

cols = [
    ("起票日", 11),
    ("Q番号", 7),
    ("枝番", 7),
    ("カテゴリ", 14),
    ("質問内容", 50),
    ("選択肢 / 補足", 30),
    ("回答(クライアント記入)", 36),
    ("回答日", 11),
    ("担当", 9),
    ("ステータス", 12),
    ("備考(社内メモ)", 26),
]
for i, (h, w) in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = HF
    c.fill = HFL
    c.alignment = WRAP_C
    c.border = BD
ws.row_dimensions[1].height = 28
ws.freeze_panes = "B2"

# データ
items = [
    # (起票日, Q, 枝, カテゴリ, 質問, 選択肢, 回答, 回答日, 担当, ステータス, 備考)
    ("2026-05-07", "Q1", "1-1", "大変更",
     "受け付けるファイル形式は Excel + PDF + Word + 画像 すべてで問題ございませんでしょうか。1ファイルあたりの容量上限はいかがいたしますでしょうか。",
     "形式: 全て可 / 一部のみ\n容量上限: ___ MB",
     "形式: 全て可\n容量上限は特には考えておりません。",
     "2026-05-07", "秋元様", "回答済", "T2-10 安全書類で実装"),
    ("2026-05-07", "Q1", "1-2", "大変更",
     "元請さまごとにフォーマットが異なるとのことでしたが、システム側でテンプレートを管理する仕組みが必要でしょうか。それともアップロード機能のみで十分でしょうか。",
     "A. テンプレート管理あり\nB. アップロードのみ",
     "元請さまの情報に記録することが可能であれば、記録し次回同様の元請さまの書類作成があった際に「前回のテンプレートを使用しますか？」など選択式としてほしいです。",
     "2026-05-07", "秋元様", "回答済", "T2-25 元請テンプレート記憶 で実装"),
    ("2026-05-07", "Q1", "1-3", "大変更",
     "入構人員情報(資格/住所/連絡先)は USER マスタから自動引用する想定でよろしいでしょうか。",
     "Yes / No",
     "Yes",
     "2026-05-07", "秋元様", "回答済", "USER マスタ拡張で対応(実装済)"),
    ("2026-05-07", "Q1", "1-4", "大変更",
     "上記のため、住所と連絡先の項目を USER マスタに新規追加してよろしいでしょうか。",
     "Yes / No",
     "Yes",
     "2026-05-07", "秋元様", "実装済", "0010 マイグレーションで追加済"),
    ("2026-05-07", "Q1", "1-5", "大変更",
     "元請さまと協力会社さまの両方それぞれに別個に提出するのでしょうか。それとも1セットで両方を兼ねるのでしょうか。",
     "A. 別個 / B. 1セット",
     "A. 別個",
     "2026-05-07", "秋元様", "回答済", "safety_documents.recipient_type で対応"),
    ("2026-05-07", "Q2", "2-1", "業務ロジック",
     "「受注済み」に該当するステータスは現状の見積書の「accepted(受注)」で合致するという理解でよろしいでしょうか。",
     "Yes / No",
     "Yes",
     "2026-05-07", "秋元様", "回答済", ""),
    ("2026-05-07", "Q2", "2-2", "業務ロジック",
     "複数の見積が1案件に紐づく場合、どの見積を概況表に反映するのが適切でしょうか。",
     "A. 最新の受注見積のみ\nB. 全受注見積の合計\nC. 1案件=1受注見積ルール",
     "B. 全受注見積の合計\n→複数の見積書が作成される場合もある\nまた、一度登録したものを修正や削除する場合もあります。",
     "2026-05-07", "秋元様", "回答済", "T2-14 で実装予定(既存サマリ修正含む)"),
    ("2026-05-07", "Q3", "3-1", "業務ロジック",
     "工事経歴書の出力形式は PDF / Excel / 両方 のうち、どれが望ましいでしょうか。",
     "PDF / Excel / 両方",
     "Excel",
     "2026-05-07", "秋元様", "回答済", "ExcelJS で実装予定"),
    ("2026-05-07", "Q3", "3-2", "業務ロジック",
     "既存のテンプレート(建業許可で使用される様式)がございましたら、ご共有いただけますでしょうか。",
     "あり(添付) / なし",
     "あり(工事経歴書のExcel書式)",
     "2026-05-07", "秋元様", "回答済", "テンプレート受領待ち"),
    ("2026-05-07", "Q3", "3-3", "業務ロジック",
     "1案件で複数回請求書を発行された場合、1案件あたり1通の工事経歴書を作成するのか、請求書ごとに作成するのか、いずれがよろしいでしょうか。",
     "A. 1案件1通\nB. 請求書ごと",
     "B. 請求書ごと\n→ただし、場合によっては合計も確認したいため、「案件ごと」or「請求ごと」を選択して出力できるとうれしいです。",
     "2026-05-07", "秋元様", "回答済", "T2-26 切替機能 を実装予定"),
    ("2026-05-07", "Q3", "3-4", "業務ロジック",
     "和暦の表記は「令和7年5月7日」形式でよろしいでしょうか。",
     "Yes / No",
     "Yes",
     "2026-05-07", "秋元様", "回答済", ""),
    ("2026-05-07", "Q4", "4-1a", "業務ロジック", "経費管理表 - 見積金額 の入力者は?",
     "自動 / 事務手入力 / 案件責任者",
     "自動\n→見積書が「ステータス:受注」となった場合、見積金額部分に反映されていくようにしたい。",
     "2026-05-07", "秋元様", "実装済", "/pc/expense で自動集計"),
    ("2026-05-07", "Q4", "4-1b", "業務ロジック", "経費管理表 - 仕入(材料)金額 の入力者は?",
     "自動 / 事務手入力 / 案件責任者",
     "自動\n→仕入先さまからいただく毎月の請求書を打ち込むなどを行う部分が別途ほしい。と思いました…。",
     "2026-05-07", "秋元様", "実装済", "T2-20 仕入先請求書モジュール完了"),
    ("2026-05-07", "Q4", "4-1c", "業務ロジック", "経費管理表 - リース金額 の入力者は?",
     "自動 / 事務手入力 / 案件責任者",
     "自動\n→上記同様",
     "2026-05-07", "秋元様", "実装済", "supplier_invoices.invoice_type=lease"),
    ("2026-05-07", "Q4", "4-1d", "業務ロジック", "経費管理表 - 協力会社(人工/残業/金額) の入力者は?",
     "自動 / 事務手入力 / 案件責任者",
     "自動\n→上記同様",
     "2026-05-07", "秋元様", "実装済", "supplier_invoices.invoice_type=subcontractor"),
    ("2026-05-07", "Q4", "4-1e", "業務ロジック", "経費管理表 - 諸経費 の入力者は?",
     "自動 / 事務手入力 / 案件責任者",
     "案件責任者または自動\n→見積書を作成した案件責任者にしかわからないため。",
     "2026-05-07", "秋元様", "実装済", "supplier_invoices.invoice_type=misc"),
    ("2026-05-07", "Q4", "4-1f", "業務ロジック", "経費管理表 - 車両交通費 の入力者は?",
     "自動(領収書) / 手入力",
     "自動(領収書)\n→提出された領収書は必ず、総務課が目視し、振分を行う。振分で車両交通費となった場合、自動で集計されてほしい。",
     "2026-05-07", "秋元様", "実装済", "receipts.category=交通費 で自動集計"),
    ("2026-05-07", "Q4", "4-1g", "業務ロジック", "経費管理表 - 人件費(人工/残業/出張費) の入力者は?",
     "自動(REPORT3) / 手入力",
     "自動(REPORT3)",
     "2026-05-07", "秋元様", "実装済", "project_cost_aggregates"),
    ("2026-05-07", "Q4", "4-1h", "業務ロジック", "経費管理表 - 宿泊費(レオパレス/ホテル) の入力者は?",
     "自動(領収書) / 手入力",
     "手入力\n→自動にしたいが、現時点ではできない。",
     "2026-05-07", "秋元様", "実装済", "現状は領収書から集計、将来的に宿泊費明細追加検討"),
    ("2026-05-07", "Q4", "4-2", "業務ロジック",
     "見積金額は見積書から自動取得する形でよろしいでしょうか。",
     "Yes / No",
     "Yes(税抜き)",
     "2026-05-07", "秋元様", "実装済", "subtotal_cents から取得"),
    ("2026-05-07", "Q4", "4-3", "業務ロジック",
     "協力会社の金額入力方法 — 別途請求書を入力する仕組みが必要でしょうか。",
     "A. 必要(機能追加)\nB. 不要(手入力で対応)",
     "A. 必要(機能追加)",
     "2026-05-07", "秋元様", "実装済", "T2-20 で仕入先請求書モジュール完了"),
    ("2026-05-07", "Q5", "5-1", "マスタ",
     "「職人スキル」の評価軸はどのようなものを想定されていますでしょうか。",
     "数値スコア / 自己申告 / 上司評価 / 複合 / その他",
     "複合\n→現在対応している人事考課評価を数値にし、自己申告と多面的などの評価を行うことで、乖離を見るのもいいと考えています。",
     "2026-05-07", "秋元様", "回答済", "user_skills テーブルで実装予定"),
    ("2026-05-07", "Q5", "5-2", "マスタ",
     "「人事評価データ」の閲覧権限はどなたまでとされますでしょうか。",
     "本人のみ / 直属の上司まで / 経営層のみ",
     "選択項目のなかで言えば、基本的には経営層まで。ただし、フィードバック内容などそのなかでも閲覧可能範囲を定めたい。",
     "2026-05-07", "秋元様", "回答済", "RLS + visible_to カラムで制御予定"),
    ("2026-05-07", "Q5", "5-3", "マスタ",
     "「経験現場履歴」は REPORT3 から自動構築する形でよろしいでしょうか。",
     "Yes / No",
     "Yes",
     "2026-05-07", "秋元様", "回答済", "T2-15 内で集計ビュー作成予定"),
    ("2026-05-07", "Q6", "6-1", "マスタ",
     "「松竹梅」の判定基準はどのように設定されますでしょうか。",
     "案件作成時に手動 / 案件規模で自動 / その他",
     "案件作成時に手動\n※社長へ要確認",
     "2026-05-07", "秋元様", "社長確認待ち", "判定基準の最終決定を社長へ"),
    ("2026-05-07", "Q6", "6-2", "マスタ",
     "役職別単価は組織図の役職テーブルから決定する形でよろしいでしょうか(例:職長=標準の120%)。",
     "Yes / No(別ロジック指定)",
     "No(別ロジック指定)\n→別途作成します。",
     "2026-05-07", "秋元様", "回答済", "別ロジック秋元様提供待ち"),
    ("2026-05-07", "Q6", "6-3", "マスタ",
     "単価決定のタイミングは、見積入力時に自動計算まで実施 か、推奨単価を表示してご担当者が調整 か、どちらが望ましいでしょうか。",
     "A. 自動計算\nB. 推奨表示+手動調整",
     "B. 推奨表示+手動調整\n※社長へ要確認",
     "2026-05-07", "秋元様", "社長確認待ち", ""),
    ("2026-05-07", "Q7", "7-1", "マスタ",
     "Excel のテンプレート様式は社長様にてご用意いただけますでしょうか。それとも当方で雛形をご提案いたしましょうか。",
     "A. 社長提供\nB. 当方で提案",
     "B. 当方で提案\n→用意も可能だが、雛形があると助かります。",
     "2026-05-07", "秋元様", "回答済", "雛形作成 → クライアント提示"),
    ("2026-05-07", "Q7", "7-2", "マスタ",
     "「公開/非公開」のフラグは組織図全体に対するものか、個別の更新ごとに切り替えるものか、いずれをご想定でしょうか。",
     "A. 全体 / B. 個別",
     "B. 個別\n→役職や部署によって制限をかけたい",
     "2026-05-07", "秋元様", "実装済", "is_visible_to_all を ORG マスタに追加済"),
    ("2026-05-07", "Q7", "7-3", "マスタ",
     "「役職履歴」の粒度は日単位 / 月単位 / 任意の日付指定 のどれが業務上扱いやすいでしょうか。",
     "日単位 / 月単位 / 任意日付",
     "月単位\n→●年●ヶ月の形式",
     "2026-05-07", "秋元様", "実装済", "user_org_history.start_year_month"),
    ("2026-05-07", "Q8", "8-*", "REPORT3 / 領収書",
     "REPORT3 の工種分類 4 階層化について。詳細は別シート「Q8-中分類」「Q8-小分類」「Q8-組合せ」「Q8-反映方針」をご参照ください。",
     "別シートに記入",
     "★社長にて対応中",
     "", "社長", "社長確認待ち", "別シート 4 枚を社長記入待ち"),
    ("2026-05-07", "Q9", "9-1", "REPORT3 / 領収書",
     "標準勤務時間の前提を 8:00〜17:00 / 休憩1時間(計8時間)としてよろしいでしょうか。",
     "Yes / No(別時間指定)",
     "No(別時間指定)\n→現場はあっているが、内作は8:00~17:30\nまた、夜勤なども可能性あるため固定化は難しいと考えます。",
     "2026-05-07", "秋元様", "回答済", "T2-01 勤怠で柔軟設計"),
    ("2026-05-07", "Q9", "9-2", "REPORT3 / 領収書",
     "夜勤の判定はどの時間帯で行うのが適切でしょうか(例:勤務開始時刻が22:00以降を夜勤)。",
     "境界時刻: ___ 時 以降",
     "境界時刻: 22 時 以降、朝5時まで",
     "2026-05-07", "秋元様", "回答済", "勤怠ロジックに反映予定"),
    ("2026-05-07", "Q9", "9-3", "REPORT3 / 領収書",
     "夜勤分は除外とするか、別カウントとするか、いずれがよろしいでしょうか。",
     "A. 除外\nB. 別カウント(夜勤手当対象)",
     "恐らくB\n→給与計算で言えば、倍率が変わるため",
     "2026-05-07", "秋元様", "回答済", "夜勤手当倍率の確認必要(別途質問)"),
    ("2026-05-07", "Q9", "9-4", "REPORT3 / 領収書",
     "スケジュール機能(配車表・人員配置)の先行実装が必要との理解でよろしいでしょうか。",
     "Yes(先行実装) / No(別途簡易実装)",
     "Yes(先行実装)",
     "2026-05-07", "秋元様", "回答済", "T3-02 スケジュールを先に実装"),
    ("2026-05-07", "Q10", "10-1", "REPORT3 / 領収書",
     "「精算の有無」とは、会社負担で立替済→精算が必要 か 個人負担(精算不要)の2択でよろしいでしょうか。",
     "Yes / No(別の意味)",
     "No(別の意味)\n→領収書の提出については、会社のガソリンカードやETCなどがあるため、会社のカードで支払いしている場合は提出のみ。自分のお金で立替を行った場合は精算対象になるため、精算申請を行うようにしてほしい。",
     "2026-05-07", "秋元様", "実装済", "T2-21 領収書フロー2系統 完了"),
    ("2026-05-07", "Q10", "10-2", "REPORT3 / 領収書",
     "カテゴリ選択肢(消耗品/食事代/交通費/宿泊費/その他)で過不足ございませんでしょうか。追加項目があればご記入ください。",
     "追加項目: ___________",
     "追加項目:駐車料金\nまた可能であれば、一部細分化してほしい。\n消耗品:事務用品or材料\n交通費→ガソリンor高速料金\n別途、食事代を選択した場合は、だれといったのかも記載or選択をさせたい(接待交際費なのか福利厚生費なのかを判断するため)",
     "2026-05-07", "秋元様", "実装済", "T2-22 カテゴリ階層化 完了"),
    ("2026-05-07", "Q10", "10-3", "REPORT3 / 領収書",
     "1件あたりの金額上限(例:5万円超は別途承認フロー)を設定する必要はございますでしょうか。",
     "A. 不要 / B. 必要(上限: ___ 円)",
     "A. 不要\n→ただし、現場名と日付を選択した上で、領収書の申請や登録を行ってもらうが、その際に申請月と領収月の差が1月以上ある場合は警告してほしい。また、年度をまたいだ場合は申請不可としてほしい。",
     "2026-05-07", "秋元様", "実装済", "T2-23 申請日制約 完了"),
    ("2026-05-07", "Q10", "10-4", "REPORT3 / 領収書",
     "通知先となる「総務課」は既存「事務」ロール内の特定スタッフ(フラグ)としてよろしいでしょうか。それとも新規ロールに分離しますか。",
     "A. 既存ロール+フラグ\nB. 新規ロール分離",
     "選択対象が理解できていないため、入力します。総務課内でも現金管理や精算対応は限られた人で対応しているため、同じ総務課でも閲覧制限や通知対象外としたい。また、別途経理課へは精算が行われました。などのおかねの動きを通知させたい。",
     "2026-05-07", "秋元様", "実装済", "is_general_affairs_member / is_accounting_member フラグ追加"),
    ("2026-05-07", "Q11", "11-1", "その他",
     "Eight 連携の方向性。Eight→顧客マスタへの取込のみで十分でしょうか。書き出しも必要でしょうか。",
     "A. 取込のみ\nB. 双方向",
     "書き出しができると完全にEightは解約できます。ただし、Eightについては絶対ツールに取り入れたいわけではありません。",
     "2026-05-07", "秋元様", "保留", "優先度低・後日検討"),
    ("2026-05-07", "Q11", "11-2", "その他",
     "Eight の API キーは社長様側でご用意いただける想定でよろしいでしょうか。",
     "Yes / No(検討中)",
     "No\n→おそらく難しいかと",
     "2026-05-07", "秋元様", "保留", "API キー入手困難なら実装不可"),
    ("2026-05-07", "Q12", "12-1", "その他",
     "事務ロールを「総務課」「施工管理課」の2つに分けるか、ロールは「事務」のままで部署マスタで識別するか、いずれがよろしいでしょうか。",
     "A. ロール分割\nB. 部署マスタで識別",
     "B. 部署マスタで識別",
     "2026-05-07", "秋元様", "実装済", "0010 で org_departments + profiles.department_id"),

    # === 新規追加 (UI/UX) ===
    ("2026-05-08", "Q13", "13-1", "UI/UX",
     "PC 画面のメニューが 25 項目に増えました。サイドバーの配置方向のご希望はいかがでしょうか。",
     "A. 縦表示(画面左に縦並び/現状)\nB. 横表示(画面上部に横並びタブ)",
     "",
     "", "", "未回答", "クライアント送付待ち"),
    ("2026-05-08", "Q14", "14-1", "UI/UX",
     "メニュー項目の整理方法のご希望はいかがでしょうか。",
     "A. すべてフラットに並べる(現状)\nB. カテゴリで見出し付き縦リスト\nC. カテゴリで開閉できるツリー形式",
     "",
     "", "", "未回答", ""),
    ("2026-05-08", "Q14", "14-2", "UI/UX",
     "下記の 6 カテゴリ案で不足や変更のご希望はございますでしょうか。"
     "\n▼業務フロー(ダッシュボード/通知/日報/承認/ヒヤリハット)"
     "\n▼営業・経理(見積/請求書/入金管理/客先別売上/領収書/仕入先請求書/経費管理表)"
     "\n▼マスタ(現場/客先/ユーザー/部署/役職/工種/単価/資格)"
     "\n▼設備管理(工具/車両)"
     "\n▼評価・人事(ランキング)"
     "\n▼個人(プロフィール)",
     "自由記述",
     "",
     "", "", "未回答", ""),
    ("2026-05-08", "Q15", "15-1", "UI/UX",
     "スマホ画面(SP)のメニューは現状のホーム画面 7 連クイックカード + ボトムナビで問題ございませんでしょうか。",
     "Yes / No(変更希望)",
     "",
     "", "", "未回答", "PC ナビとは別管理"),

    # === 後続予定の質問プレースホルダ(将来追加用) ===
    ("", "Q16", "", "", "(空欄 — 次の質問追加時にここから記入)", "", "", "", "", "未回答", ""),
    ("", "Q17", "", "", "", "", "", "", "", "未回答", ""),
    ("", "Q18", "", "", "", "", "", "", "", "未回答", ""),
    ("", "Q19", "", "", "", "", "", "", "", "未回答", ""),
    ("", "Q20", "", "", "", "", "", "", "", "未回答", ""),
]

for row_idx, item in enumerate(items, 2):
    for col_idx, val in enumerate(item, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = JF
        c.alignment = WRAP if col_idx in (5, 6, 7, 11) else WRAP_C
        c.border = BD
    cat = item[3]
    status = item[9]
    if cat in CAT_FILL:
        ws.cell(row=row_idx, column=4).fill = CAT_FILL[cat]
    if status in STATUS_FILL:
        ws.cell(row=row_idx, column=10).fill = STATUS_FILL[status]
    # 回答欄を黄色く
    ws.cell(row=row_idx, column=7).fill = ANSWER_FILL
    # 行高
    qlen = sum(len(str(v) or "") for v in item)
    ws.row_dimensions[row_idx].height = max(38, min(140, qlen * 0.32))

# ========================================
# Q8 サブシート(既存質問票と同じ構造)
# ========================================
def add_q8_chubun():
    ws2 = wb.create_sheet("Q8-中分類")
    for i, w in enumerate([16, 8, 35], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.merge_cells("A1:C1")
    ws2["A1"] = "Q8 中分類リスト(大分類ごとに記入)/ ★ 社長対応中"
    ws2["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 24
    ws2.merge_cells("A2:C2")
    ws2["A2"] = "「大分類」を選択した時に表示される「中分類」のリストをご記入ください。「配管」は記載済(参考)。"
    ws2["A2"].alignment = WRAP
    ws2["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws2.row_dimensions[2].height = 30
    for i, h in enumerate(["大分類", "No.", "中分類(ご記入欄)"], 1):
        c = ws2.cell(row=4, column=i, value=h)
        c.font = HF
        c.fill = HFL
        c.alignment = WRAP_C
        c.border = BD
    chubun = {
        "配管": ["冷媒", "冷温水", "冷却水", "排水", "食品プロセス", "ケミカルプロセス", "エネルギープロセス", "蒸気", "エアー"],
        "鳶": [""] * 10, "保温": [""] * 10, "溶接": [""] * 10, "雑工事": [""] * 10,
    }
    row = 5
    for d, lst in chubun.items():
        start = row
        for j, item in enumerate(lst, 1):
            c1 = ws2.cell(row=row, column=1, value=d if j == 1 else "")
            c2 = ws2.cell(row=row, column=2, value=j)
            c3 = ws2.cell(row=row, column=3, value=item)
            for c in [c1, c2, c3]:
                c.font = JF
                c.border = BD
            c1.font = JFB
            c1.alignment = Alignment(vertical="center", horizontal="center")
            c2.alignment = WRAP_C
            c3.alignment = WRAP
            c3.fill = PatternFill("solid", fgColor="EEEEEE") if d == "配管" else ANSWER_FILL
            ws2.row_dimensions[row].height = 22
            row += 1
        if len(lst) > 1:
            ws2.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        row += 1


def add_q8_shoubun():
    ws3 = wb.create_sheet("Q8-小分類")
    for i, w in enumerate([18, 8, 35], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Q8 小分類リスト(中分類ごとに記入)/ ★ 社長対応中"
    ws3["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 24
    ws3.merge_cells("A2:C2")
    ws3["A2"] = "「中分類」を選択した時に表示される「小分類」のリストをご記入ください。「排水」は記載済(参考)。"
    ws3["A2"].alignment = WRAP
    ws3["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws3.row_dimensions[2].height = 30
    for i, h in enumerate(["中分類", "No.", "小分類(ご記入欄)"], 1):
        c = ws3.cell(row=4, column=i, value=h)
        c.font = HF
        c.fill = HFL
        c.alignment = WRAP_C
        c.border = BD
    shoubun = {
        "冷媒": [""] * 10, "冷温水": [""] * 10, "冷却水": [""] * 10,
        "排水": ["塩ビ/のり付け", "塩ビ/溶接", "塩ビ/加工管・フランジ止め", "SGPW/溶接", "SGPW/ネジ", "SGPW/加工管・フランジ閉め", "SUS", "SGP_黒"],
        "食品プロセス": [""] * 10, "ケミカルプロセス": [""] * 10, "エネルギープロセス": [""] * 10,
        "蒸気": [""] * 10, "エアー": [""] * 10,
    }
    row = 5
    for cn, lst in shoubun.items():
        start = row
        for j, item in enumerate(lst, 1):
            c1 = ws3.cell(row=row, column=1, value=cn if j == 1 else "")
            c2 = ws3.cell(row=row, column=2, value=j)
            c3 = ws3.cell(row=row, column=3, value=item)
            for c in [c1, c2, c3]:
                c.font = JF
                c.border = BD
            c1.font = JFB
            c1.alignment = Alignment(vertical="center", horizontal="center")
            c2.alignment = WRAP_C
            c3.alignment = WRAP
            c3.fill = PatternFill("solid", fgColor="EEEEEE") if cn == "排水" else ANSWER_FILL
            ws3.row_dimensions[row].height = 22
            row += 1
        if len(lst) > 1:
            ws3.merge_cells(start_row=start, start_column=1, end_row=row - 1, end_column=1)
        row += 1


def add_q8_combo():
    ws4 = wb.create_sheet("Q8-組合せ")
    ws4.column_dimensions["A"].width = 22
    for i in range(2, 7):
        ws4.column_dimensions[get_column_letter(i)].width = 14
    ws4.merge_cells("A1:F1")
    ws4["A1"] = "Q8 特大分類 × 大分類 の組み合わせ制限 / ★ 社長対応中"
    ws4["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 24
    ws4.merge_cells("A2:F2")
    ws4["A2"] = "業務上発生しうる組合せに「○」、発生しない/制限する組合せに「×」をご記入ください。空白は判断保留。"
    ws4["A2"].alignment = WRAP
    ws4["A2"].font = Font(name="Yu Gothic", size=9, color="666666")
    ws4.row_dimensions[2].height = 30
    ws4.cell(row=4, column=1, value="特大分類 \\ 大分類")
    for i, d in enumerate(["配管", "鳶", "保温", "溶接", "雑工事"], 2):
        ws4.cell(row=4, column=i, value=d)
    for i in range(1, 7):
        c = ws4.cell(row=4, column=i)
        c.font = HF
        c.fill = HFL
        c.alignment = WRAP_C
        c.border = BD
    ws4.row_dimensions[4].height = 24
    for r, t in enumerate(["冷凍冷蔵", "ケミカルプラント", "大型半導体", "食品工場"], 5):
        c = ws4.cell(row=r, column=1, value=t)
        c.font = JFB
        c.alignment = WRAP_C
        c.fill = PatternFill("solid", fgColor="DDEEFF")
        c.border = BD
        for col in range(2, 7):
            cc = ws4.cell(row=r, column=col, value="")
            cc.fill = ANSWER_FILL
            cc.border = BD
            cc.alignment = WRAP_C
            cc.font = JF
        ws4.row_dimensions[r].height = 24


def add_q8_policy():
    ws5 = wb.create_sheet("Q8-反映方針")
    ws5.column_dimensions["A"].width = 50
    ws5.column_dimensions["B"].width = 30
    ws5.merge_cells("A1:B1")
    ws5["A1"] = "Q8 既存システムへの反映方針 / ★ 社長対応中"
    ws5["A1"].font = Font(name="Yu Gothic", size=14, bold=True, color="1A3A6A")
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 24
    ws5.cell(row=3, column=1, value="質問")
    ws5.cell(row=3, column=2, value="ご回答")
    for i in range(1, 3):
        c = ws5.cell(row=3, column=i)
        c.font = HF
        c.fill = HFL
        c.alignment = WRAP_C
        c.border = BD
    q4 = [
        "現状システムには3階層のシードデータ(33件)が登録されています。新4階層体系の導入方針はいかがいたしますでしょうか。\n\n選択肢: A. 既存3階層を完全削除→新4階層に入れ替え / B. 既存に追記する形で4階層を加える",
        "過去日報の移行対応は不要との認識でよろしいでしょうか。(現時点で本番データはない想定です)\n\n選択肢: Yes(不要) / No(必要)",
    ]
    for i, q in enumerate(q4, 4):
        ws5.cell(row=i, column=1, value=q)
        ws5.cell(row=i, column=2, value="")
        for j in range(1, 3):
            c = ws5.cell(row=i, column=j)
            c.font = JF
            c.alignment = WRAP
            c.border = BD
        ws5.cell(row=i, column=2).fill = ANSWER_FILL
        ws5.row_dimensions[i].height = 100


add_q8_chubun()
add_q8_shoubun()
add_q8_combo()
add_q8_policy()

# ========================================
# シート: 更新履歴
# ========================================
ws_log = wb.create_sheet("更新履歴")
ws_log.column_dimensions["A"].width = 14
ws_log.column_dimensions["B"].width = 16
ws_log.column_dimensions["C"].width = 80

for i, h in enumerate(["日付", "担当", "内容"], 1):
    c = ws_log.cell(row=1, column=i, value=h)
    c.font = HF
    c.fill = HFL
    c.alignment = WRAP_C
    c.border = BD
ws_log.row_dimensions[1].height = 22

logs = [
    ("2026-05-07", "AI(匠)", "Q1〜Q12 初版作成、クライアント送付"),
    ("2026-05-07", "秋元様", "Q1〜Q7・Q9〜Q12 回答記入、社長対応分(Q6-1, Q6-3, Q8)を保留"),
    ("2026-05-07", "AI(匠)", "回答を Master Backlog へ反映、T2-20〜T2-27 として 8 項目追加"),
    ("2026-05-07", "AI(匠)", "実装着手分を「実装済」に更新、本番デプロイ完了(コミット 26a4c50)"),
    ("2026-05-08", "AI(匠)", "UI/UX 関連の Q13〜Q15 を新規追記(メニュー縦/横・整理方法・SP 確認)"),
    ("", "", ""),
    ("", "", ""),
    ("", "", ""),
]
for i, (d, who, msg) in enumerate(logs, 2):
    ws_log.cell(row=i, column=1, value=d).font = JF
    ws_log.cell(row=i, column=2, value=who).font = JF
    ws_log.cell(row=i, column=3, value=msg).font = JF
    for col in [1, 2, 3]:
        c = ws_log.cell(row=i, column=col)
        c.alignment = WRAP
        c.border = BD
    ws_log.row_dimensions[i].height = 24

out_path = r"C:\Users\liim1\Desktop\SAKURAOS_ヒアリング項目.xlsx"
wb.save(out_path)
print("Saved:", out_path)
